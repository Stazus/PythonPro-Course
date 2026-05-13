"""
Prosty arkusz kalkulacyjny w Pythonie z użyciem openpyxl.

Ten program tworzy plik Excel o nazwie 'finanse.xlsx', w którym:
- w kolumnie A umieszczone są nazwy wydatków (np. "Czynsz", "Jedzenie"),
- w kolumnie B umieszczone są odpowiadające wartości wydatków,
- w komórce pomiżej wartości obliczana jest suma wszystkich wydatków za pomocą formuły Excela.
Dodatkowo program ustawia szerokośc kolumn i wypisuje zawartość arkusza w terminalu.
Żeby program działał należy zainstalować bibliotekę openpyxl - .... 
w środowisku wirtualnym: python3 -m venv venv i je aktywować: source venv/bin/activate
"""

from openpyxl import Workbook

# Tworzymy nowy skoroszyt i wybieramy aktywny arkusz
wb = Workbook()
ws = wb.active
ws.title = "Finanse"

# Przykładowe wydatki
wydatki = [
    ("Czynsz", 1200),
    ("Jedzenie", 800),
    ("Transport", 150)
]

# Wstawiamy dane do arkusza w kolumnach A i B
for i, (nazwa, kwota) in enumerate(wydatki, start=1):
    ws[f"A{i}"] = nazwa # Nazwa wydatku w kolumnie A
    ws[f"B{i}"] = kwota # Wartość wydatku w kolumnie B
    
# Dodajemy formułę sumy w komórce poniżej ostatniego wydatku
ws[f"B{len(wydatki)+1}"] = f"=SUM(B1:B{len(wydatki)})"

# Ustawiamy szerokość kolumn, aby dane były widoczne w Excelu
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 10

# Zapisujemy arkusz do pliku
wb.save("finanse.xlsx")
print("Plik finanse.xlsx został utworzony.")

# Wyświetlamy zawartość arkusza w terminalu
for row in ws.iter_rows(values_only=True):
    print(row)